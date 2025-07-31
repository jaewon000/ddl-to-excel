# ddl_to_excel_converter.py
import re
import pandas as pd
import os


import re

def parse_ddl_tables(ddl_string):
    # 1단계: 전체 테이블 구조 파싱
    pattern = r'CREATE TABLE `(.+?)` \((.*?)\)\s*ENGINE=.*?(?:COMMENT=\'(.+?)\')?;'
    table_definitions = []
    
    matches = re.finditer(pattern, ddl_string, re.DOTALL)
    
    for match in matches:
        table_name = match.group(1)
        columns_section = match.group(2).strip()
        table_comment = match.group(3) if match.group(3) else ""
        
        # 2단계: 컬럼 정보 파싱
        columns = parse_columns(columns_section)
        
        table_definitions.append({
            'table_name': table_name,
            'columns': columns,
            'table_comment': table_comment
        })
    
    return table_definitions

def parse_columns(columns_section):
    columns = []
    
    # PRIMARY KEY, UNIQUE KEY 등 제약조건 제거하고 컬럼만 추출
    lines = columns_section.split('\n')
    
    for line in lines:
        line = line.strip()
        if not line or line.endswith(','):
            line = line.rstrip(',')
        # 제약조건 스킵
        if (line.startswith('PRIMARY KEY') or 
            line.startswith('UNIQUE KEY') or 
            line.startswith('KEY') or
            line.startswith('INDEX') or
            line.startswith('CONSTRAINT')):
            continue
        # 컬럼 정의 파싱
        if line.startswith('`'):
            # NULL/NOT NULL 정보도 추출
            nullable = 'NULL'
            if 'NOT NULL' in line:
                nullable = 'NOT NULL'
            column_info = parse_single_column(line)
            if column_info:
                column_info['nullable'] = nullable
                columns.append(column_info)
    
    return columns

def parse_single_column(line):
    # 컬럼명 추출
    column_name_match = re.match(r'`(.+?)`', line)
    if not column_name_match:
        return None
    
    column_name = column_name_match.group(1)
    
    # 컬럼 타입 추출
    type_match = re.search(r'`\w+`\s+(\w+(?:\(\d+(?:,\d+)?\))?(?:\s+\w+)*)', line)
    column_type = type_match.group(1) if type_match else ""
    
    # 컬럼 코멘트 추출
    comment_match = re.search(r"COMMENT '(.+?)'", line)
    column_comment = comment_match.group(1) if comment_match else ""
    
    return {
        'column_name': column_name,
        'column_type': column_type,
        'comment': column_comment
    }



def parse_ddl_to_excel(ddl_string, output_filename="table_definition.xlsx"):
    """
    DDL 문자열을 파싱하여 테이블 정의서 엑셀 파일로 저장합니다.

    :param ddl_string: 데이터 정의 언어(DDL) 전체 텍스트
    :param output_filename: 저장할 엑셀 파일 이름
    """
    
    # 새 파싱 로직에 맞는 엑셀 저장 로직
    table_definitions = parse_ddl_tables(ddl_string)

    if not table_definitions:
        print("DDL에서 테이블 정의를 찾을 수 없습니다.")
        return

    # 상세 시트용 데이터 (기존 레이아웃 유지)
    detail_rows = []
    for table in table_definitions:
        table_name = table['table_name']
        table_comment = table.get('table_comment', '')
        table_name_with_comment = f"{table_name} ({table_comment})" if table_comment else table_name
        columns = table['columns']
        # PK/UNIQUE/FK 컬럼명 추출
        columns_section = None
        import re
        table_pattern = rf'CREATE TABLE `{re.escape(table_name)}` \((.*?)\)\s*ENGINE=.*?(?:COMMENT=\'.*?\')?;'
        m_table = re.search(table_pattern, ddl_string, re.DOTALL)
    raise NotImplementedError("이 함수는 멀티파일 지원을 위해 main에서만 호출하세요.")
    # 요약 시트 데이터 (기존 레이아웃 유지)
    summary_rows = []
    for table in table_definitions:
        table_name = table['table_name']
        table_comment = table.get('table_comment', '')
        table_name_with_comment = f"{table_name} ({table_comment})" if table_comment else table_name
        summary_rows.append({
            '주제영역': '',
            '테이블ID': table_name,
            '테이블명': table_name_with_comment,
            '테이블설명': table_comment
        })
    summary_df = pd.DataFrame(summary_rows, columns=['주제영역', '테이블ID', '테이블명', '테이블설명'])

    # 엑셀 저장 및 스타일 (기존 스타일 최대한 유지)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            # 요약 시트
            summary_df.to_excel(writer, index=False, sheet_name='테이블목록')
            # 상세 시트
            detail_df.to_excel(writer, index=False, sheet_name='테이블정의서')
            workbook = writer.book
            ws_summary = writer.sheets['테이블목록']
            ws_detail = writer.sheets['테이블정의서']
            # 헤더 스타일
            header_font = Font(bold=True, color='000000')
            header_fill = PatternFill(fill_type='solid', fgColor='EFEFEF')
            header_align = Alignment(horizontal='center', vertical='center')
            for ws in [ws_summary, ws_detail]:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                # 컬럼 너비 자동조정
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2
                # 전체 셀에 얇은 테두리
                thin = Side(border_style="thin", color="B0B0B0")
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                # 헤더(1행)만 굵은 박스 테두리(요약시트, 상세시트 모두)
                thick = Side(border_style="thick", color="000000")
                for col in range(1, ws.max_column+1):
                    cell = ws.cell(row=1, column=col)
                    top = thick
                    bottom = thick
                    left = thick if col == 1 else cell.border.left
                    right = thick if col == ws.max_column else cell.border.right
                    cell.border = Border(top=top, left=left, right=right, bottom=bottom)
            # 상세시트: 테이블명 세로 병합 및 테이블별 굵은 하단 테두리
            ws = ws_detail
            table_col_idx = None
            for col in range(1, ws.max_column+1):
                if ws.cell(row=1, column=col).value == '테이블명':
                    table_col_idx = col
                    break
            if table_col_idx:
                prev_val = None
                merge_start = None
                thick = Side(border_style="thick", color="000000")
                for row_idx in range(2, ws.max_row+2):  # 1-based, 1은 헤더
                    val = ws.cell(row=row_idx, column=table_col_idx).value if row_idx <= ws.max_row else None
                    if val != prev_val:
                        # 병합
                        if merge_start and row_idx-merge_start > 1:
                            ws.merge_cells(start_row=merge_start, start_column=table_col_idx, end_row=row_idx-1, end_column=table_col_idx)
                            cell = ws.cell(row=merge_start, column=table_col_idx)
                            cell.alignment = Alignment(vertical='center', horizontal='center')
                        # 박스 테두리(이전 테이블 전체 구간)
                        if prev_val is not None and merge_start and row_idx-merge_start > 0:
                            for r in range(merge_start, row_idx):
                                for c in range(1, ws.max_column+1):
                                    cell = ws.cell(row=r, column=c)
                                    # 굵은 상단/하단/좌/우 테두리 적용
                                    top = thick if r == merge_start else cell.border.top
                                    bottom = thick if r == row_idx-1 else cell.border.bottom
                                    left = thick if c == 1 else cell.border.left
                                    right = thick if c == ws.max_column else cell.border.right
                                    cell.border = Border(top=top, left=left, right=right, bottom=bottom)
                        merge_start = row_idx
                        prev_val = val
        print(f"✅ 성공! 테이블 정의서가 '{os.path.abspath(output_filename)}' 파일로 저장되었습니다.")
    except Exception as e:
        print(f"❌ 에러: 엑셀 파일 저장에 실패했습니다. ({e})")


# --- 스크립트 실행 ---
if __name__ == '__main__':
    import glob
    import os
    import pandas as pd
    from collections import OrderedDict
    import re
    # sql 폴더 내 모든 .sql 파일 읽기
    sql_dir = os.path.join(os.path.dirname(__file__), 'sql')
    sql_files = sorted([f for f in glob.glob(os.path.join(sql_dir, '*.sql'))])
    if not sql_files:
        print(f"❌ sql 폴더에 .sql 파일이 없습니다.")
        exit(1)

    # 각 파일별로 DDL 파싱
    all_table_defs = OrderedDict()  # {주제영역(파일명): [table_def,...]}
    all_ddl = dict()  # {주제영역(파일명): ddl_string}
    for sql_path in sql_files:
        with open(sql_path, encoding='utf-8') as f:
            ddl = f.read()
        file_key = os.path.splitext(os.path.basename(sql_path))[0]  # 예: am, mm
        table_defs = parse_ddl_tables(ddl)
        if table_defs:
            all_table_defs[file_key] = table_defs
            all_ddl[file_key] = ddl

    # 전체 테이블 목록(요약) 시트 데이터
    summary_rows = []
    for topic, tables in all_table_defs.items():
        for table in tables:
            table_name = table['table_name']
            table_comment = table.get('table_comment', '')
            table_name_with_comment = f"{table_name} ({table_comment})" if table_comment else table_name
            summary_rows.append({
                '주제영역': topic,
                '테이블ID': table_name,
                '테이블명': table_name_with_comment,
                '테이블설명': table_comment
            })
    summary_df = pd.DataFrame(summary_rows, columns=['주제영역', '테이블ID', '테이블명', '테이블설명'])

    # 각 파일별 상세 시트 데이터
    detail_dfs = OrderedDict()
    for topic, tables in all_table_defs.items():
        ddl = all_ddl[topic]
        detail_rows = []
        for table in tables:
            table_name = table['table_name']
            table_comment = table.get('table_comment', '')
            table_name_with_comment = f"{table_name} ({table_comment})" if table_comment else table_name
            columns = table['columns']
            # PK/UNIQUE/FK 컬럼명 추출
            columns_section = None
            table_pattern = rf'CREATE TABLE `{re.escape(table_name)}` \((.*?)\)\s*ENGINE=.*?(?:COMMENT=\'.*?\')?;'
            m_table = re.search(table_pattern, ddl, re.DOTALL)
            pk_cols = []
            unique_keys = []
            fk_infos = []
            if m_table:
                columns_section = m_table.group(1)
                # PK
                pk_match = re.search(r'PRIMARY KEY \(([^\)]+)\)', columns_section)
                if pk_match:
                    pk_cols = [col.strip('` ') for col in pk_match.group(1).split(',')]
                # UNIQUE KEY
                unique_keys = re.findall(r'UNIQUE KEY `(.+?)` \(([^\)]+)\)', columns_section)
                # FK
                fk_matches = re.findall(r'CONSTRAINT `(.+?)` FOREIGN KEY \(([^\)]+)\) REFERENCES `(.+?)` \(([^\)]+)\)', columns_section)
                for fk_name, fk_cols_str, ref_table, ref_cols_str in fk_matches:
                    fk_cols = [col.strip('` ') for col in fk_cols_str.split(',')]
                    ref_cols = [col.strip('` ') for col in ref_cols_str.split(',')]
                    fk_infos.append((fk_name, fk_cols, ref_table, ref_cols))
            col_no = 1
            for col in columns:
                # 데이터 타입과 길이 분리
                col_type = col['column_type']
                col_type_main = col_type
                col_length = ''
                m = re.match(r'([a-zA-Z]+)\((\d+)\)', col_type)
                if m:
                    col_type_main = m.group(1)
                    col_length = m.group(2)
                # NULL/PK/기본값/인덱스 정보
                is_null = 'N' if col.get('nullable', '') == 'NOT NULL' else 'Y'
                is_pk = 'Y' if col['column_name'] in pk_cols else 'N'
                # 기본값 추출
                default_val = ''
                default_pattern = rf'`{re.escape(col["column_name"])}.*?DEFAULT\s+((?:NULL)|(?:\'[^"]*\')|(?:[\w\d\.\-]+))'
                default_match = None
                if columns_section:
                    for line in columns_section.split('\n'):
                        if line.strip().startswith(f'`{col["column_name"]}`'):
                            default_match = re.search(default_pattern, line.strip())
                            break
                if default_match:
                    default_val = default_match.group(1)
                    if default_val.startswith("'") and default_val.endswith("'"):
                        default_val = default_val[1:-1]
                elif is_null == 'Y':
                    default_val = 'NULL'
                detail_rows.append({
                    '테이블명': table_name_with_comment,
                    'No': col_no,
                    '컬럼 논리명': col['comment'] if col['comment'] else col['column_name'],
                    '컬럼 물리명': col['column_name'],
                    '데이터 타입': col_type_main,
                    '길이': col_length,
                    'NULL': is_null,
                    'PK': is_pk,
                    '기본값': default_val,
                    '코멘트 / 인덱스 정보': col['comment']
                })
                col_no += 1
            # 인덱스/제약조건 정보 행 추가
            # PK
            if pk_cols:
                detail_rows.append({
                    '테이블명': table_name_with_comment,
                    'No': '',
                    '컬럼 논리명': '**Index**',
                    '컬럼 물리명': f"PRIMARY ({', '.join(pk_cols)})",
                    '데이터 타입': '',
                    '길이': '',
                    'NULL': '',
                    'PK': '',
                    '기본값': '',
                    '코멘트 / 인덱스 정보': 'UNIQUE'
                })
            # UNIQUE KEY
            for key_name, key_cols_str in unique_keys:
                key_cols = [col.strip('` ') for col in key_cols_str.split(',')]
                detail_rows.append({
                    '테이블명': table_name_with_comment,
                    'No': '',
                    '컬럼 논리명': '**Index**',
                    '컬럼 물리명': f"{key_name} ({', '.join(key_cols)})",
                    '데이터 타입': '',
                    '길이': '',
                    'NULL': '',
                    'PK': '',
                    '기본값': '',
                    '코멘트 / 인덱스 정보': 'UNIQUE'
                })
            # FOREIGN KEY
            for fk_name, fk_cols, ref_table, ref_cols in fk_infos:
                detail_rows.append({
                    '테이블명': table_name_with_comment,
                    'No': '',
                    '컬럼 논리명': '**ForeignKey**',
                    '컬럼 물리명': f"{fk_name} ({', '.join(fk_cols)}) -> {ref_table}({', '.join(ref_cols)})",
                    '데이터 타입': '',
                    '길이': '',
                    'NULL': '',
                    'PK': '',
                    '기본값': '',
                    '코멘트 / 인덱스 정보': 'FOREIGN KEY'
                })
            # 테이블 간 구분 빈 행
            if columns:
                detail_rows.append({k: '' for k in ['테이블명','No','컬럼 논리명','컬럼 물리명','데이터 타입','길이','NULL','PK','기본값','코멘트 / 인덱스 정보']})
        detail_df = pd.DataFrame(detail_rows, columns=['테이블명', 'No', '컬럼 논리명', '컬럼 물리명', '데이터 타입', '길이', 'NULL', 'PK', '기본값', '코멘트 / 인덱스 정보'])
        detail_dfs[topic] = detail_df

    # 엑셀 저장 및 스타일 (기존 스타일 최대한 유지)
    output_filename = "table_definition.xlsx"
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            # 요약 시트(공통)
            summary_df.to_excel(writer, index=False, sheet_name='테이블목록')
            # 각 주제영역별 상세 시트
            for topic, detail_df in detail_dfs.items():
                detail_df.to_excel(writer, index=False, sheet_name=topic)
            workbook = writer.book
            ws_summary = writer.sheets['테이블목록']
            # 요약 시트 스타일
            header_font = Font(bold=True, color='000000')
            header_fill = PatternFill(fill_type='solid', fgColor='EFEFEF')
            header_align = Alignment(horizontal='center', vertical='center')
            for ws in [ws_summary] + [writer.sheets[topic] for topic in detail_dfs.keys()]:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                # 컬럼 너비 자동조정
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2
                # 전체 셀에 얇은 테두리
                thin = Side(border_style="thin", color="B0B0B0")
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                # 헤더(1행)만 굵은 박스 테두리
                thick = Side(border_style="thick", color="000000")
                for col in range(1, ws.max_column+1):
                    cell = ws.cell(row=1, column=col)
                    top = thick
                    bottom = thick
                    left = thick if col == 1 else cell.border.left
                    right = thick if col == ws.max_column else cell.border.right
                    cell.border = Border(top=top, left=left, right=right, bottom=bottom)

            # --- 요약시트(테이블목록) 주제영역 병합 및 굵은 박스 테두리 ---
            ws = ws_summary
            # '주제영역' 컬럼 인덱스 찾기
            topic_col_idx = None
            for col in range(1, ws.max_column+1):
                if ws.cell(row=1, column=col).value == '주제영역':
                    topic_col_idx = col
                    break
            if topic_col_idx:
                prev_val = None
                merge_start = None
                thick = Side(border_style="thick", color="000000")
                for row_idx in range(2, ws.max_row+2):  # 1-based, 1은 헤더
                    val = ws.cell(row=row_idx, column=topic_col_idx).value if row_idx <= ws.max_row else None
                    if val != prev_val:
                        # 병합
                        if merge_start and row_idx-merge_start > 1:
                            ws.merge_cells(start_row=merge_start, start_column=topic_col_idx, end_row=row_idx-1, end_column=topic_col_idx)
                            cell = ws.cell(row=merge_start, column=topic_col_idx)
                            cell.alignment = Alignment(vertical='center', horizontal='center')
                        # 박스 테두리(이전 주제영역 전체 구간)
                        if prev_val is not None and merge_start and row_idx-merge_start > 0:
                            for r in range(merge_start, row_idx):
                                for c in range(1, ws.max_column+1):
                                    cell = ws.cell(row=r, column=c)
                                    # 굵은 상단/하단/좌/우 테두리 적용
                                    top = thick if r == merge_start else cell.border.top
                                    bottom = thick if r == row_idx-1 else cell.border.bottom
                                    left = thick if c == 1 else cell.border.left
                                    right = thick if c == ws.max_column else cell.border.right
                                    cell.border = Border(top=top, left=left, right=right, bottom=bottom)
                        merge_start = row_idx
                        prev_val = val
            # 상세시트: 테이블명 세로 병합 및 테이블별 굵은 하단 테두리 (각 상세시트별 반복)
            for topic in detail_dfs.keys():
                ws = writer.sheets[topic]
                table_col_idx = None
                for col in range(1, ws.max_column+1):
                    if ws.cell(row=1, column=col).value == '테이블명':
                        table_col_idx = col
                        break
                if table_col_idx:
                    prev_val = None
                    merge_start = None
                    thick = Side(border_style="thick", color="000000")
                    for row_idx in range(2, ws.max_row+2):  # 1-based, 1은 헤더
                        val = ws.cell(row=row_idx, column=table_col_idx).value if row_idx <= ws.max_row else None
                        if val != prev_val:
                            # 병합
                            if merge_start and row_idx-merge_start > 1:
                                ws.merge_cells(start_row=merge_start, start_column=table_col_idx, end_row=row_idx-1, end_column=table_col_idx)
                                cell = ws.cell(row=merge_start, column=table_col_idx)
                                cell.alignment = Alignment(vertical='center', horizontal='center')
                            # 박스 테두리(이전 테이블 전체 구간)
                            if prev_val is not None and merge_start and row_idx-merge_start > 0:
                                for r in range(merge_start, row_idx):
                                    for c in range(1, ws.max_column+1):
                                        cell = ws.cell(row=r, column=c)
                                        # 굵은 상단/하단/좌/우 테두리 적용
                                        top = thick if r == merge_start else cell.border.top
                                        bottom = thick if r == row_idx-1 else cell.border.bottom
                                        left = thick if c == 1 else cell.border.left
                                        right = thick if c == ws.max_column else cell.border.right
                                        cell.border = Border(top=top, left=left, right=right, bottom=bottom)
                            merge_start = row_idx
                            prev_val = val
        print(f"✅ 성공! 테이블 정의서가 '{os.path.abspath(output_filename)}' 파일로 저장되었습니다.")
    except Exception as e:
        print(f"❌ 에러: 엑셀 파일 저장에 실패했습니다. ({e})")
